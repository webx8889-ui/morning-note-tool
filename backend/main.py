# backend/main.py
from fastapi import FastAPI, UploadFile, File
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from io import BytesIO
from openpyxl import load_workbook
from datetime import date, datetime


app = FastAPI(title="Morning Note HTML Generator")

origins = [
    "http://127.0.0.1:5500",
    "http://localhost:5500",
    "http://127.0.0.1:8000",
    "https://webxds.com",
    "https://www.webxds.com", 
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ---------- Helper functions ----------

def cell(ws, addr):
    """Read a cell value from worksheet, strip if string."""
    v = ws[addr].value
    return v.strip() if isinstance(v, str) else (str(v) if v is not None else "")


def fmt_number(value, decimals=2):
    if value is None or value == "":
        return ""
    try:
        num = float(str(value).replace(",", ""))
        fmt = f"{{:,.{decimals}f}}"
        return fmt.format(num)
    except (ValueError, TypeError):
        return str(value)


def fmt_percent(value, decimals=2, show_sign=True):
    if value is None or value == "":
        return ""
    s = str(value).strip()
    try:
        if s.endswith("%"):
            num = float(s.replace("%", "").replace(",", ""))
        else:
            num = float(s.replace(",", ""))
            if abs(num) < 1:
                num = num * 100
        sign = "+" if show_sign and num > 0 else ""
        fmt = f"{{:.{decimals}f}}"
        return f"{sign}{fmt.format(num)}%"
    except (ValueError, TypeError):
        return s


def perc_color(value_str):
    s = str(value_str).strip()
    if not s:
        return "#6b7280"
    if s.startswith("-"):
        return "#b91c1c"
    if s.startswith("+") or (s and s[0].isdigit()):
        return "#15803d"
    return "#6b7280"


# ---------- HTML Template ----------

HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1"/>
<title>{title}</title>
</head>
<body style="margin:0; padding:0; background-color:#f3f4f6; font-family: system-ui, -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif; color:#111827;">

<table width="100%" cellpadding="0" cellspacing="0" border="0" style="background-color:#f3f4f6;">
  <tr>
    <td align="center" valign="top">
      <table width="100%" cellpadding="0" cellspacing="0" border="0" style="max-width:1024px; padding:24px 16px 40px;">
        <tr>
          <td>

            <!-- HEADER -->
            <table width="100%" cellpadding="0" cellspacing="0" border="0">
              <tr>
                <td align="left" valign="middle" style="font-size:14px; font-weight:600; color:#4b5563;">{date_line}</td>
                <td align="right" valign="middle">
                  <a href="{podcast_link}">
                    <img src="https://res.cloudinary.com/dqhtbzeuj/image/upload/v1772081294/Bajaj_broking_logo_fbos5n.jpg" alt="Bajaj Broking" style="width:150px; height:auto;"/>
                  </a>
                </td>
              </tr>
            </table>

            <!-- MAIN TITLE -->
            <table width="100%" cellpadding="0" cellspacing="0" border="0">
              <tr>
                <td style="font-size:28px; font-weight:700; color:#0f172a; padding:16px 0 16px 0;">{main_heading}</td>
              </tr>
            </table>

            <!-- MARKET SNAPSHOT CARD -->
            <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#ffffff; border-radius:12px; box-shadow:0 8px 20px rgba(15,23,42,0.06); margin-bottom:16px;">
              <tr>
                <td style="padding:16px 18px;">
                  <table width="100%" cellpadding="0" cellspacing="0" border="0">
                    <tr>
                      <td style="font-size:18px; font-weight:700; color:#0f172a; padding:0 0 12px 0;">{market_snapshot_heading}</td>
                    </tr>
                  </table>

                  <!-- ROW 1 -->
                  <table width="100%" cellpadding="0" cellspacing="0" border="0">
                    <tr>
                      <td valign="top" width="33.33%" style="padding:0 8px 12px 0;">
                        <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#f9fafb; border-radius:10px;">
                          <tr><td style="padding:10px 12px;">
                            <div style="font-size:11px; text-transform:uppercase; letter-spacing:0.05em; color:#6b7280; margin-bottom:4px;">{gift_label}</div>
                            <div style="font-size:16px; font-weight:600; color:#111827;">{gift_value}</div>
                            <div style="font-size:11px; margin-top:2px; color:{gift_color}; font-weight:500;">{gift_change}</div>
                          </td></tr>
                        </table>
                      </td>
                      <td valign="top" width="33.33%" style="padding:0 8px 12px 0;">
                        <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#f9fafb; border-radius:10px;">
                          <tr><td style="padding:10px 12px;">
                            <div style="font-size:11px; text-transform:uppercase; letter-spacing:0.05em; color:#6b7280; margin-bottom:4px;">{nifty_label}</div>
                            <div style="font-size:16px; font-weight:600; color:#111827;">{nifty_value}</div>
                            <div style="font-size:11px; margin-top:2px; color:{nifty_color}; font-weight:500;">{nifty_change}</div>
                          </td></tr>
                        </table>
                      </td>
                      <td valign="top" width="33.33%" style="padding:0 0 12px 0;">
                        <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#f9fafb; border-radius:10px;">
                          <tr><td style="padding:10px 12px;">
                            <div style="font-size:11px; text-transform:uppercase; letter-spacing:0.05em; color:#6b7280; margin-bottom:4px;">{sensex_label}</div>
                            <div style="font-size:16px; font-weight:600; color:#111827;">{sensex_value}</div>
                            <div style="font-size:11px; margin-top:2px; color:{sensex_color}; font-weight:500;">{sensex_change}</div>
                          </td></tr>
                        </table>
                      </td>
                    </tr>

                    <!-- ROW 2 -->
                    <tr>
                      <td valign="top" width="33.33%" style="padding:0 8px 12px 0;">
                        <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#f9fafb; border-radius:10px;">
                          <tr><td style="padding:10px 12px;">
                            <div style="font-size:11px; text-transform:uppercase; letter-spacing:0.05em; color:#6b7280; margin-bottom:4px;">{bank_label}</div>
                            <div style="font-size:16px; font-weight:600; color:#111827;">{bank_value}</div>
                            <div style="font-size:11px; margin-top:2px; color:{bank_color}; font-weight:500;">{bank_change}</div>
                          </td></tr>
                        </table>
                      </td>
                      <td valign="top" width="33.33%" style="padding:0 8px 12px 0;">
                        <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#f9fafb; border-radius:10px;">
                          <tr><td style="padding:10px 12px;">
                            <div style="font-size:11px; text-transform:uppercase; letter-spacing:0.05em; color:#6b7280; margin-bottom:4px;">{vix_label}</div>
                            <div style="font-size:16px; font-weight:600; color:#111827;">{vix_value}</div>
                            <div style="font-size:11px; margin-top:2px; color:{vix_color}; font-weight:500;">{vix_change}</div>
                          </td></tr>
                        </table>
                      </td>
                      <td valign="top" width="33.33%" style="padding:0 0 12px 0;">
                        <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#f9fafb; border-radius:10px;">
                          <tr><td style="padding:10px 12px;">
                            <div style="font-size:11px; text-transform:uppercase; letter-spacing:0.05em; color:#6b7280; margin-bottom:4px;">{usdinr_label}</div>
                            <div style="font-size:16px; font-weight:600; color:#111827;">{usdinr_value}</div>
                            <div style="font-size:11px; margin-top:2px; color:{usdinr_color}; font-weight:500;">{usdinr_change}</div>
                          </td></tr>
                        </table>
                      </td>
                    </tr>
                  </table>

                  <!-- PODCAST SECTION -->
                  <table width="100%" cellpadding="0" cellspacing="0" border="0" style="margin-top:12px;">
                    <tr>
                      <td>
                        <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#0f172a; border-radius:12px; color:#e5e7eb;">
                          <tr>
                            <td style="padding:16px 18px;">
                              <table cellpadding="0" cellspacing="0" border="0" style="margin-bottom:8px;">
                                <tr>
                                  <td style="padding:4px 10px; border-radius:999px; background:#1f2937; color:#e5e7eb; font-size:11px; font-weight:500;">{podcast_tagline}</td>
                                </tr>
                              </table>
                              <p style="font-size:12px; margin:8px 0 0; line-height:1.5; color:#e5e7eb;">{podcast_para1}</p>
                              <table cellpadding="0" cellspacing="0" border="0" style="margin-top:8px;">
                                <tr>
                                  <td align="center" style="background:#005DAC; border-radius:999px;">
                                    <a href="{podcast_link}" style="display:block; padding:8px 16px; font-size:12px; font-weight:600; color:#facc15; text-decoration:none;">🎧 LISTEN NOW 🎧</a>
                                  </td>
                                </tr>
                              </table>
                              <div style="font-size:11px; margin-top:6px; color:#9ca3af;">{podcast_footer}</div>
                            </td>
                          </tr>
                        </table>
                      </td>
                    </tr>
                  </table>

                </td>
              </tr>
            </table>

            <!-- INDIA MARKET RECAP -->
            <table width="100%" cellpadding="0" cellspacing="0" border="0" style="margin-top:24px;">
              <tr><td style="font-size:18px; font-weight:700; color:#0f172a; padding-bottom:12px;">{recap_heading}</td></tr>
            </table>
            <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#ffffff; border-radius:12px; box-shadow:0 8px 20px rgba(15,23,42,0.06); margin-bottom:16px;">
              <tr>
                <td style="padding:16px 18px;">
                  <p style="font-size:14px; line-height:1.6; margin:0 0 12px 0;">{recap_para1}</p>
                  <p style="font-size:14px; line-height:1.6; margin:0 0 12px 0;">{recap_para2}</p>
                  <p style="font-size:14px; line-height:1.6; margin:0 0 12px 0;">{recap_para3}</p>
                  <p style="font-size:11px; color:#6b7280; margin:0;">{recap_source}</p>
                </td>
              </tr>
            </table>

            <!-- TRADING PLAYBOOK -->
            <table width="100%" cellpadding="0" cellspacing="0" border="0" style="margin-top:24px;">
              <tr><td style="font-size:18px; font-weight:700; color:#0f172a; padding-bottom:12px;">{playbook_heading}</td></tr>
            </table>
            <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#ffffff; border-radius:12px; box-shadow:0 8px 20px rgba(15,23,42,0.06); margin-bottom:16px;">
              <tr>
                <td style="padding:16px 18px;">
                  <table width="100%" cellpadding="0" cellspacing="0" border="0">
                    <tr><td style="font-size:14px; font-weight:600; color:#374151; padding-bottom:4px;">{flows_heading}</td></tr>
                    <tr><td style="font-size:11px; color:#6b7280; padding-bottom:10px;">{flows_subtext}</td></tr>
                  </table>

                  <!-- FII/DII TABLE -->
                  <table width="100%" cellpadding="0" cellspacing="0" border="0" style="font-size:12px; border-collapse:collapse;">
                    <thead>
                      <tr>
                        <th align="left" style="padding:8px 6px; background:#eff6ff; border-bottom:1px solid #e5e7eb; font-size:11px; text-transform:uppercase; letter-spacing:0.04em; color:#111827;">Participant</th>
                        <th align="left" style="padding:8px 6px; background:#eff6ff; border-bottom:1px solid #e5e7eb; font-size:11px; text-transform:uppercase; letter-spacing:0.04em; color:#111827;">Previous Day (&#8377; Cr)</th>
                        <th align="left" style="padding:8px 6px; background:#eff6ff; border-bottom:1px solid #e5e7eb; font-size:11px; text-transform:uppercase; letter-spacing:0.04em; color:#111827;">MTD (&#8377; Cr)</th>
                        <th align="left" style="padding:8px 6px; background:#eff6ff; border-bottom:1px solid #e5e7eb; font-size:11px; text-transform:uppercase; letter-spacing:0.04em; color:#111827;">YTD (&#8377; Cr)</th>
                      </tr>
                    </thead>
                    <tbody>
                      <tr>
                        <td style="padding:8px 6px; border-bottom:1px solid #e5e7eb;">FII</td>
                        <td style="padding:8px 6px; border-bottom:1px solid #e5e7eb; color:{fii_prev_color}; font-weight:500;">{fii_prev}</td>
                        <td style="padding:8px 6px; border-bottom:1px solid #e5e7eb; color:{fii_mtd_color}; font-weight:500;">{fii_mtd}</td>
                        <td style="padding:8px 6px; border-bottom:1px solid #e5e7eb; color:{fii_ytd_color}; font-weight:500;">{fii_ytd}</td>
                      </tr>
                      <tr>
                        <td style="padding:8px 6px; border-bottom:1px solid #e5e7eb;">DII</td>
                        <td style="padding:8px 6px; border-bottom:1px solid #e5e7eb; color:{dii_prev_color}; font-weight:500;">{dii_prev}</td>
                        <td style="padding:8px 6px; border-bottom:1px solid #e5e7eb; color:{dii_mtd_color}; font-weight:500;">{dii_mtd}</td>
                        <td style="padding:8px 6px; border-bottom:1px solid #e5e7eb; color:{dii_ytd_color}; font-weight:500;">{dii_ytd}</td>
                      </tr>
                    </tbody>
                  </table>
                  <p style="font-size:11px; color:#6b7280; margin:8px 0 0 0;">{flows_source}</p>

                  <!-- RANGE LEVELS -->
                  <p style="font-size:11px; color:#6b7280; margin:16px 0 4px 0;"><strong>{range_heading}</strong></p>
                  <table width="100%" cellpadding="0" cellspacing="0" border="0" style="font-size:12px; border-collapse:collapse;">
                    <thead>
                      <tr>
                        <th align="left" style="padding:8px 6px; background:#eff6ff; border-bottom:1px solid #e5e7eb; font-size:11px; text-transform:uppercase; letter-spacing:0.04em; color:#111827;">Index</th>
                        <th align="left" style="padding:8px 6px; background:#eff6ff; border-bottom:1px solid #e5e7eb; font-size:11px; text-transform:uppercase; letter-spacing:0.04em; color:#111827;">Support 1</th>
                        <th align="left" style="padding:8px 6px; background:#eff6ff; border-bottom:1px solid #e5e7eb; font-size:11px; text-transform:uppercase; letter-spacing:0.04em; color:#111827;">Support 2</th>
                        <th align="left" style="padding:8px 6px; background:#eff6ff; border-bottom:1px solid #e5e7eb; font-size:11px; text-transform:uppercase; letter-spacing:0.04em; color:#111827;">Resistance 1</th>
                        <th align="left" style="padding:8px 6px; background:#eff6ff; border-bottom:1px solid #e5e7eb; font-size:11px; text-transform:uppercase; letter-spacing:0.04em; color:#111827;">Resistance 2</th>
                      </tr>
                    </thead>
                    <tbody>
                      <tr>
                        <td style="padding:8px 6px; border-bottom:1px solid #e5e7eb;">{range_row1_index}</td>
                        <td style="padding:8px 6px; border-bottom:1px solid #e5e7eb;">{range_row1_s1}</td>
                        <td style="padding:8px 6px; border-bottom:1px solid #e5e7eb;">{range_row1_s2}</td>
                        <td style="padding:8px 6px; border-bottom:1px solid #e5e7eb;">{range_row1_r1}</td>
                        <td style="padding:8px 6px; border-bottom:1px solid #e5e7eb;">{range_row1_r2}</td>
                      </tr>
                      <tr>
                        <td style="padding:8px 6px; border-bottom:1px solid #e5e7eb;">{range_row2_index}</td>
                        <td style="padding:8px 6px; border-bottom:1px solid #e5e7eb;">{range_row2_s1}</td>
                        <td style="padding:8px 6px; border-bottom:1px solid #e5e7eb;">{range_row2_s2}</td>
                        <td style="padding:8px 6px; border-bottom:1px solid #e5e7eb;">{range_row2_r1}</td>
                        <td style="padding:8px 6px; border-bottom:1px solid #e5e7eb;">{range_row2_r2}</td>
                      </tr>
                    </tbody>
                  </table>
                  <p style="font-size:11px; color:#6b7280; margin:8px 0 0 0;">{range_comment}</p>

                </td>
              </tr>
            </table>

            <!-- GLOBAL MARKET PULSE -->
            <table width="100%" cellpadding="0" cellspacing="0" border="0" style="margin-top:24px;">
              <tr><td style="font-size:18px; font-weight:700; color:#0f172a; padding-bottom:12px;">{global_heading}</td></tr>
            </table>
            <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#ffffff; border-radius:12px; box-shadow:0 8px 20px rgba(15,23,42,0.06); margin-bottom:16px;">
              <tr>
                <td style="padding:16px 18px;">
                  <p style="font-size:14px; line-height:1.6; margin:0 0 10px 0;">{global_para1}</p> 
                  <p style="font-size:11px; color:#6b7280; margin:0;">{global_source}</p>
                </td>
              </tr>
            </table>

            <!-- STOCKS IN FOCUS -->
            <table width="100%" cellpadding="0" cellspacing="0" border="0" style="margin-top:24px;">
              <tr><td style="font-size:18px; font-weight:700; color:#0f172a; padding-bottom:12px;">{stocks_heading}</td></tr>
            </table>
            <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#ffffff; border-radius:12px; box-shadow:0 8px 20px rgba(15,23,42,0.06); margin-bottom:16px;">
              <tr>
                <td style="padding:16px 18px;">
                  <table width="100%" cellpadding="0" cellspacing="0" border="0" style="font-size:12px; border-collapse:collapse;">
                    <thead>
                      <tr>
                        <th align="left" style="padding:8px 6px; background:#eff6ff; border-bottom:1px solid #e5e7eb; font-size:11px; text-transform:uppercase; letter-spacing:0.04em; color:#111827;">Symbol</th>
                        <th align="left" style="padding:8px 6px; background:#eff6ff; border-bottom:1px solid #e5e7eb; font-size:11px; text-transform:uppercase; letter-spacing:0.04em; color:#111827;">Price %</th>
                        <th align="left" style="padding:8px 6px; background:#eff6ff; border-bottom:1px solid #e5e7eb; font-size:11px; text-transform:uppercase; letter-spacing:0.04em; color:#111827;">OI %</th>
                        <th align="left" style="padding:8px 6px; background:#eff6ff; border-bottom:1px solid #e5e7eb; font-size:11px; text-transform:uppercase; letter-spacing:0.04em; color:#111827;">Interpretation</th>
                      </tr>
                    </thead>
                    <tbody>{stocks_rows_html}</tbody>
                  </table>
                  <p style="font-size:11px; color:#6b7280; margin:8px 0 0 0;">{stocks_source}</p>
                </td>
              </tr>
            </table>

            <!-- CORPORATE HIGHLIGHTS -->
            <table width="100%" cellpadding="0" cellspacing="0" border="0" style="margin-top:24px;">
              <tr><td style="font-size:18px; font-weight:700; color:#0f172a; padding-bottom:12px;">{corp_heading}</td></tr>
            </table>
            <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#ffffff; border-radius:12px; box-shadow:0 8px 20px rgba(15,23,42,0.06);">
              <tr>
                <td style="padding:16px 18px;">
                  <table width="100%" cellpadding="0" cellspacing="0" border="0" style="font-size:12px; border-collapse:collapse;">
                    <thead>
                      <tr>
                        <th align="left" style="padding:8px 6px; background:#eff6ff; border-bottom:1px solid #e5e7eb; font-size:11px; text-transform:uppercase; letter-spacing:0.04em; color:#111827;">Company</th>
                        <th align="left" style="padding:8px 6px; background:#eff6ff; border-bottom:1px solid #e5e7eb; font-size:11px; text-transform:uppercase; letter-spacing:0.04em; color:#111827;">Key News / Impact</th>
                      </tr>
                    </thead>
                    <tbody>{corp_rows_html}</tbody>
                  </table>
                  <p style="font-size:11px; color:#6b7280; margin:8px 0 0 0;">{corp_source}</p>
                </td>
              </tr>
            </table>

            <!-- DISCLAIMER -->
            <table width="100%" cellpadding="0" cellspacing="0" border="0" style="margin-top:24px;">
              <tr>
                <td style="padding:0px; font-size:11px; color:#374151; line-height:1.6;">
                  <p style="margin:0 0 8px 0;"><strong>Note:</strong> {note_text}</p>
                  <p style="margin:0 0 8px 0;"><strong>Disclaimer:</strong> Please do not reply to this email, as responses to this address will not be received or monitored. For any queries, feel free to contact us at <a href="mailto:connect@bajajbroking.in" style="color:#1414DF; font-weight:450;">connect@bajajbroking.in</a>.</p>
                  <p style="margin:0 0 8px 0; line-height:16px;">Investments in the securities market are subject to market risk. Read all related documents carefully before investing. REG OFFICE: Bajaj Auto Limited Complex, Mumbai Pune Road, Akurdi, Pune 411035. Corp. Office: Bajaj Financial Securities Limited. 1st Floor, Mantri IT Park, Tower B, Unit No. 9 &amp; 10, Viman Nagar, Pune, Maharashtra 411014. SEBI Registration No.: INZ000218931 | BSE Cash/F&amp;O/CDS Member ID: 6706 | NSE Cash/F&amp;O/CDS Member ID: 90177 | SEBI DP Registration No.: IN-DP-418-2019 | CDSL DP No.: 12088600 | NSDL DP No.: IN304030 | AMFI Registration No.: ARN-163403 | SEBI Registration No. (Research Analyst/Entity): INH000010043 | Website: <a href="https://www.bajajbroking.in" style="color:#1414DF; font-weight:450;">https://www.bajajbroking.in</a></p>
                  <p style="margin:0 0 8px 0;">Compliance Officer: Mr. Boudhayan Ghosh (For Broking/DP/Research) | Email: <a href="mailto:compliance_sec@bajajbroking.in" style="color:#1414DF; font-weight:450;">compliance_sec@bajajbroking.in</a> | Contact No.: 020-4857 4486.</p>
                  <p style="margin:0 0 8px 0;">For any queries/information, you can call the toll-free number 1800-833-8888 or write to us on <a href="mailto:connect@bajajbroking.in" style="color:#1414DF; font-weight:450;">connect@bajajbroking.in</a>. For any Investor Grievances, write to compliance on <a href="mailto:compliance_sec@bajajbroking.in" style="color:#1414DF; font-weight:450;">compliance_sec@bajajbroking.in</a> or <a href="mailto:compliance_dp@bajajbroking.in" style="color:#1414DF; font-weight:450;">compliance_dp@bajajbroking.in</a></p>
                  <p style="margin:0;">Kindly refer to <a href="https://www.bajajbroking.in/disclaimer" style="color:#1414DF; font-weight:450;">https://www.bajajbroking.in/disclaimer</a> for detailed disclaimer and risk factors.</p>
                </td>
              </tr>
            </table>

          </td>
        </tr>
      </table>
    </td>
  </tr>
</table>
</body>
</html>"""


# ---------- Core generator ----------

def generate_html_from_excel(excel_bytes: bytes) -> str:
    wb = load_workbook(BytesIO(excel_bytes), data_only=True)
    ws = wb["Sheet1"]

    # ── MARKET SNAPSHOT ──────────────────────────────────────────
    raw_gift_value    = cell(ws, "D9")
    raw_gift_change   = cell(ws, "D10")
    raw_nifty_value   = cell(ws, "G9")
    raw_nifty_change  = cell(ws, "G10")
    raw_sensex_value  = cell(ws, "J9")
    raw_sensex_change = cell(ws, "J10")

    raw_bank_value    = cell(ws, "D13")
    raw_bank_change   = cell(ws, "D14")
    raw_vix_value     = cell(ws, "G13")
    raw_vix_change    = cell(ws, "G14")
    raw_usdinr_value  = cell(ws, "J13")
    raw_usdinr_change = cell(ws, "J14")

    gift_value    = fmt_number(raw_gift_value,   1)
    gift_change   = fmt_percent(raw_gift_change,  2)
    nifty_value   = fmt_number(raw_nifty_value,  2)
    nifty_change  = fmt_percent(raw_nifty_change, 2)
    sensex_value  = fmt_number(raw_sensex_value,  2)
    sensex_change = fmt_percent(raw_sensex_change,2)
    bank_value    = fmt_number(raw_bank_value,   2)
    bank_change   = fmt_percent(raw_bank_change,  2)
    vix_value     = fmt_number(raw_vix_value,    2)
    vix_change    = fmt_percent(raw_vix_change,   2)
    usdinr_value  = fmt_number(raw_usdinr_value,  2)
    usdinr_change = fmt_percent(raw_usdinr_change,2)

    # ── FII / DII ─────────────────────────────────────────────────
    raw_fii_prev = cell(ws, "F33")
    raw_fii_mtd  = cell(ws, "H33")
    raw_fii_ytd  = cell(ws, "J33")
    raw_dii_prev = cell(ws, "F34")
    raw_dii_mtd  = cell(ws, "H34")
    raw_dii_ytd  = cell(ws, "J34")

    fii_prev = fmt_number(raw_fii_prev, 2) or "—"
    fii_mtd  = fmt_number(raw_fii_mtd,  2) or "—"
    fii_ytd  = fmt_number(raw_fii_ytd,  2) or "—"
    dii_prev = fmt_number(raw_dii_prev, 2) or "—"
    dii_mtd  = fmt_number(raw_dii_mtd,  2) or "—"
    dii_ytd  = fmt_number(raw_dii_ytd,  2) or "—"

    # ── STOCKS IN FOCUS (Long/Short Build-up) ────────────────────
    # Excel rows 49–52: D=Symbol, F=Price%, H=OI%, J=Interpretation
    stocks_rows_html = ""
    for r in range(49, 54):
        symbol = cell(ws, f"D{r}")
        raw_price = cell(ws, f"F{r}")
        raw_oi    = cell(ws, f"H{r}")
        interp    = cell(ws, f"J{r}")

        if not symbol:
            continue

        price = fmt_percent(raw_price, 2, show_sign=True)
        oi    = fmt_percent(raw_oi,    2, show_sign=True)
        p_color = perc_color(price)
        o_color = perc_color(oi)

        stocks_rows_html += f"""
      <tr>
        <td style="padding:8px 6px; border-bottom:1px solid #e5e7eb;">{symbol}</td>
        <td style="padding:8px 6px; border-bottom:1px solid #e5e7eb; color:{p_color}; font-weight:500;">{price}</td>
        <td style="padding:8px 6px; border-bottom:1px solid #e5e7eb; color:{o_color}; font-weight:500;">{oi}</td>
        <td style="padding:8px 6px; border-bottom:1px solid #e5e7eb;">{interp}</td>
      </tr>"""

    # ── CORPORATE / STOCKS IN NEWS ───────────────────────────────
    # Excel rows 56–65: D=Company, H=Update
    corp_rows_html = ""
    for r in range(57, 61):
        company = cell(ws, f"D{r}")
        update  = cell(ws, f"G{r}")
        if not company:
            continue
        corp_rows_html += f"""
      <tr>
        <td style="padding:8px 6px; border-bottom:1px solid #e5e7eb;"><strong>{company}</strong></td>
        <td style="padding:8px 6px; border-bottom:1px solid #e5e7eb;">{update}</td>
      </tr>"""

    # ── BUILD FULL CONTEXT ────────────────────────────────────────
    today_str = datetime.today().strftime("%A, %d %B %Y")

    context = {
        # Header
        "title":                  f"Bajaj Broking Morning Note – {datetime.today().strftime('%d %B %Y')}",
        "date_line":              cell(ws, "D2") or today_str,
        "main_heading":           cell(ws, "D3") or "Bajaj Broking Morning Note",

        # Market Snapshot
        "market_snapshot_heading": cell(ws, "D6") or "Market Snapshot",
        "gift_label":             cell(ws, "D8")  or "GIFT Nifty*",
        "nifty_label":            cell(ws, "G8")  or "Nifty 50*",
        "sensex_label":           cell(ws, "J8")  or "Sensex^",
        "bank_label":             cell(ws, "D12") or "Bank Nifty*",
        "vix_label":              cell(ws, "G12") or "India VIX*",
        "usdinr_label":           cell(ws, "J12") or "USDINR*",

        "gift_value":   gift_value,   "gift_change":   gift_change,   "gift_color":   perc_color(gift_change),
        "nifty_value":  nifty_value,  "nifty_change":  nifty_change,  "nifty_color":  perc_color(nifty_change),
        "sensex_value": sensex_value, "sensex_change": sensex_change, "sensex_color": perc_color(sensex_change),
        "bank_value":   bank_value,   "bank_change":   bank_change,   "bank_color":   perc_color(bank_change),
        "vix_value":    vix_value,    "vix_change":    vix_change,    "vix_color":    perc_color(vix_change),
        "usdinr_value": usdinr_value, "usdinr_change": usdinr_change, "usdinr_color": perc_color(usdinr_change),

        # Podcast
        "podcast_tagline": cell(ws, "D18") or "Today's 3-min Podcast",
        "podcast_para1":   cell(ws, "D19"),
        "podcast_link":    cell(ws, "D22") or "https://open.spotify.com/show/4T3szhxvlaEuiMpryokCGo",
        "podcast_footer":  cell(ws, "D23") or "",

        # India Market Recap
        "recap_heading": cell(ws, "D24") or "India Market Recap",
        "recap_para1":   cell(ws, "D25"),
        "recap_para2":   cell(ws, "D26") or "",
        "recap_para3":   cell(ws, "D27") or "",
        "recap_source":  cell(ws, "D28") or f"Source: Bajaj Broking Research Desk ({datetime.today().strftime('%d %b %Y')}).",

        # Trading Playbook
        "playbook_heading": cell(ws, "D30") or "Trading Playbook",
        "flows_heading":    cell(ws, "D31") or "FII/DII Flows (Cash)",
        "flows_subtext":    cell(ws, "D32") or "Previous Day, MTD & YTD",
        "fii_prev": fii_prev, "fii_prev_color": perc_color(raw_fii_prev),
        "fii_mtd":  fii_mtd,  "fii_mtd_color":  perc_color(raw_fii_mtd),
        "fii_ytd":  fii_ytd,  "fii_ytd_color":  perc_color(raw_fii_ytd),
        "dii_prev": dii_prev, "dii_prev_color": perc_color(raw_dii_prev),
        "dii_mtd":  dii_mtd,  "dii_mtd_color":  perc_color(raw_dii_mtd),
        "dii_ytd":  dii_ytd,  "dii_ytd_color":  perc_color(raw_dii_ytd),
        "flows_source":  cell(ws, "D35") or "Source: NSE / Bajaj Broking Research.",

        # Range Table
        "range_heading":    cell(ws, "D36") or "Range to Track – Key Index Levels",
        "range_row1_index": cell(ws, "D38"),
        "range_row1_s1":    fmt_number(cell(ws, "E38"), 0),
        "range_row1_s2":    fmt_number(cell(ws, "G38"), 0),
        "range_row1_r1":    fmt_number(cell(ws, "I38"), 0),
        "range_row1_r2":    fmt_number(cell(ws, "K38"), 0),
        "range_row2_index": cell(ws, "D39"),
        "range_row2_s1":    fmt_number(cell(ws, "E39"), 0),
        "range_row2_s2":    fmt_number(cell(ws, "G39"), 0),
        "range_row2_r1":    fmt_number(cell(ws, "I39"), 0),
        "range_row2_r2":    fmt_number(cell(ws, "K39"), 0),
        "range_comment":    cell(ws, "D40") or "Source: Bajaj Broking Technical Research Desk.",

        # Global Market Pulse
        "global_heading": cell(ws, "D41") or "Global Market Pulse",
        "global_para1":   cell(ws, "D42"),
        
        "global_source":  cell(ws, "D46") or "Source: Reuters, Bloomberg.",

        # Stocks in Focus
        "stocks_heading":   cell(ws, "D47") or "Stocks in Focus",
        "stocks_rows_html": stocks_rows_html,
        "stocks_source":    cell(ws, "D53") or "Source: NSE F&O Data, Bajaj Broking Research.",

        # Corporate Highlights
        "corp_heading":   cell(ws, "D54") or "Stocks in News",
        "corp_rows_html": corp_rows_html,
        "corp_source":    cell(ws, "D66") or "Source: BSE/NSE Corporate Filings.",

        # Disclaimer
        "note_text": cell(ws, "D68") or (
            f"This Morning Note is prepared using information verified by the Bajaj Broking research team "
            f"dated {datetime.today().strftime('%d/%m/%Y')} and other publicly available sources. "
            "It is meant for informational purposes only and does not constitute investment advice."
        ),
    }

    wb.close()
    return HTML_TEMPLATE.format(**context)


# ---------- FastAPI endpoint ----------

@app.post("/generate-html")
async def generate_html(file: UploadFile = File(...)):
    excel_bytes = await file.read()
    html_str = generate_html_from_excel(excel_bytes)

    buf = BytesIO(html_str.encode("utf-8"))
    filename = f"MorningNote_{date.today():%d-%b-%Y}.html"

    return StreamingResponse(
        buf,
        media_type="text/html",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
