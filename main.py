# backend/main.py
from fastapi import FastAPI, UploadFile, File
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from io import BytesIO

from openpyxl import load_workbook
from datetime import date, datetime


app = FastAPI(title="Morning Note HTML Generator")

origins = [
    "http://127.0.0.1:5500",
    "http://localhost:5500",
    "http://127.0.0.1:8000",
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ---------- Helper functions ----------

def cell(ws, addr):
    v = ws[addr].value
    return v.strip() if isinstance(v, str) else (v or "")

def fmt_number(value, decimals=2):
    if value is None or value == "":
        return ""
    try:
        num = float(str(value).replace(",", ""))
        fmt = f"{{:,.{decimals}f}}"
        return fmt.format(num)
    except ValueError:
        return str(value)

def fmt_percent(value, decimals=2, show_sign=True):
    if value is None or value == "":
        return ""
    s = str(value).strip()
    try:
        if s.endswith("%"):
            num = float(s.replace("%", "").replace(",", ""))
        else:
            num = float(s.replace(",", ""))
            if abs(num) < 1:
                num = num * 100
        sign = "+" if show_sign and num > 0 else ""
        fmt = f"{{:.{decimals}f}}"
        return f"{sign}{fmt.format(num)}%"
    except ValueError:
        return s

def perc_color(value_str):
    s = str(value_str).strip()
    if not s:
        return "#111827"
    if s.startswith("-"):
        return "#b91c1c"
    return "#15803d"

def fmt_date(v):
    if isinstance(v, datetime):
        return v.strftime("%d %b %Y")
    elif isinstance(v, (int, float)):
        return str(v)
    elif isinstance(v, str):
        for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d %b %Y"):
            try:
                d = datetime.strptime(v.strip(), fmt)
                return d.strftime("%d %b %Y")
            except ValueError:
                continue
        return v
    return ""

# ---------- HTML template (unchanged, trimmed here for brevity) ----------

HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1"/>
<title>{title}</title>
</head>
<body style="margin:0; padding:0; background-color:#f3f4f6; font-family: system-ui, -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif; color:#111827;">

<table width="100%" cellpadding="0" cellspacing="0" border="0" style="background-color:#f3f4f6;">
  <tr>
    <td align="center" valign="top">
        
      <!-- PAGE WRAPPER -->
      <table width="100%" cellpadding="0" cellspacing="0" border="0" style="max-width:1024px; padding:24px 16px 40px;">
        <tr>
          <td>

            <!-- DATE LINE -->
            <table width="100%" cellpadding="0" cellspacing="0" border="0">
              <tr>
                <!-- DATE LINE (Left) -->
                <td align="left" valign="middle" style="font-size:14px; font-weight:600; color:#4b5563;">
                  {date_line}
                </td>
                <!-- LOGO (Right) -->
                <td align="right" valign="middle">
                    <a href="#"><img src="images/bajaj-broking-logo.png"
                            alt="" style="width: 90px;height: auto;"/></a>
                </td>
              </tr>
            </table>

            <!-- H1 TITLE -->
            <table width="100%" cellpadding="0" cellspacing="0" border="0">
              <tr>
                <td style="font-size:28px; font-weight:700; color:#0f172a; padding:0 0 16px 0;">
                  {main_heading}
                </td>
              </tr>
            </table>

            <!-- MARKET SNAPSHOT CARD -->
<table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#ffffff; border-radius:12px; box-shadow:0 8px 20px rgba(15,23,42,0.06); margin-bottom:16px;">
  <tr>
    <td style="padding:16px 18px;">

      <!-- Card Heading -->
      <table width="100%" cellpadding="0" cellspacing="0" border="0">
        <tr>
          <td style="font-size:18px; font-weight:700; color:#0f172a; padding:0 0 12px 0;">
            {market_snapshot_heading}
          </td>
        </tr>
      </table>

      <!-- METRICS SECTION (Full Width on Mobile) -->
      <table width="100%" cellpadding="0" cellspacing="0" border="0">
        <tr>
          <td>
            <!-- Metrics Grid (3x2) -->
            <table width="100%" cellpadding="0" cellspacing="0" border="0">
              <tr>
                <!-- Gift Nifty -->
                <td valign="top" width="33.33%" style="padding:0 8px 12px 0;">
                  <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#f9fafb; border-radius:10px;">
                    <tr>
                      <td style="padding:10px 12px;">
                        <div style="font-size:11px; text-transform:uppercase; letter-spacing:0.05em; color:#6b7280; margin-bottom:4px;">{gift_label}</div>
                        <div style="font-size:16px; font-weight:600; color:#111827;">{gift_value}</div>
                        <div style="font-size:11px; margin-top:2px; color:{gift_color}; font-weight:500;">{gift_change}</div>
                      </td>
                    </tr>
                  </table>
                </td>

                <!-- Nifty 50 -->
                <td valign="top" width="33.33%" style="padding:0 8px 12px 0;">
                  <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#f9fafb; border-radius:10px;">
                    <tr>
                      <td style="padding:10px 12px;">
                        <div style="font-size:11px; text-transform:uppercase; letter-spacing:0.05em; color:#6b7280; margin-bottom:4px;">{nifty_label}</div>
                        <div style="font-size:16px; font-weight:600; color:#111827;">{nifty_value}</div>
                        <div style="font-size:11px; margin-top:2px; color:{nifty_color}; font-weight:500;">{nifty_change}</div>
                      </td>
                    </tr>
                  </table>
                </td>

                <!-- Sensex -->
                <td valign="top" width="33.33%" style="padding:0 0 12px 0;">
                  <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#f9fafb; border-radius:10px;">
                    <tr>
                      <td style="padding:10px 12px;">
                        <div style="font-size:11px; text-transform:uppercase; letter-spacing:0.05em; color:#6b7280; margin-bottom:4px;">{sensex_label}</div>
                        <div style="font-size:16px; font-weight:600; color:#111827;">{sensex_value}</div>
                        <div style="font-size:11px; margin-top:2px; color:{sensex_color}; font-weight:500;">{sensex_change}</div>
                      </td>
                    </tr>
                  </table>
                </td>
              </tr>

              <tr>
                <!-- Bank Nifty -->
                <td valign="top" width="33.33%" style="padding:0 8px 12px 0;">
                  <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#f9fafb; border-radius:10px;">
                    <tr>
                      <td style="padding:10px 12px;">
                        <div style="font-size:11px; text-transform:uppercase; letter-spacing:0.05em; color:#6b7280; margin-bottom:4px;">{bank_label}</div>
                        <div style="font-size:16px; font-weight:600; color:#111827;">{bank_value}</div>
                        <div style="font-size:11px; margin-top:2px; color:{bank_color}; font-weight:500;">{bank_change}</div>
                      </td>
                    </tr>
                  </table>
                </td>

                <!-- India VIX -->
                <td valign="top" width="33.33%" style="padding:0 8px 12px 0;">
                  <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#f9fafb; border-radius:10px;">
                    <tr>
                      <td style="padding:10px 12px;">
                        <div style="font-size:11px; text-transform:uppercase; letter-spacing:0.05em; color:#6b7280; margin-bottom:4px;">{vix_label}</div>
                        <div style="font-size:16px; font-weight:600; color:#111827;">{vix_value}</div>
                        <div style="font-size:11px; margin-top:2px; color:{vix_color}; font-weight:500;">{vix_change}</div>
                      </td>
                    </tr>
                  </table>
                </td>

                <!-- USDINR -->
                <td valign="top" width="33.33%" style="padding:0 0 12px 0;">
                  <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#f9fafb; border-radius:10px;">
                    <tr>
                      <td style="padding:10px 12px;">
                        <div style="font-size:11px; text-transform:uppercase; letter-spacing:0.05em; color:#6b7280; margin-bottom:4px;">{usdinr_label}</div>
                        <div style="font-size:16px; font-weight:600; color:#111827;">{usdinr_value}</div>
                        <div style="font-size:11px; margin-top:2px; color:{usdinr_color}; font-weight:500;">{usdinr_change}</div>
                      </td>
                    </tr>
                  </table>
                </td>
              </tr>
            </table>
          </td>
        </tr>
      </table>

      <!-- PODCAST SECTION -->
      <table width="100%" cellpadding="0" cellspacing="0" border="0" style="margin-top:12px;">
        <tr>
          <td>
            <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#0f172a; border-radius:12px; color:#e5e7eb;">
              <tr>
                <td style="padding:16px 18px;">

                  <!-- Pill -->
                  <table cellpadding="0" cellspacing="0" border="0" style="margin-bottom:8px;">
                    <tr>
                      <td style="padding:4px 10px; border-radius:999px; background:#1f2937; color:#e5e7eb; font-size:11px; font-weight:500;">
                        {podcast_tagline}
                      </td>
                    </tr>
                  </table>

                  <p style="font-size:12px; margin:8px 0 0; line-height:1.5;">
                    {podcast_para1}
                  </p>

                  <p style="font-size:11px; color:#e5e7eb; opacity:0.9; margin:6px 0 0; line-height:1.5;">
                    {podcast_para2}
                  </p>

                  <!-- Button -->
                  <table cellpadding="0" cellspacing="0" border="0" style="margin-top:8px;">
                    <tr>
                      <td align="center" style="background:#005DAC; border-radius:999px;">
                        <a href="{podcast_link}" style="display:block; padding:8px 16px; font-size:12px; font-weight:600; color:#facc15; text-decoration:none;">
                          🎧 LISTEN NOW 🎧
                        </a>
                      </td>
                    </tr>
                  </table>

                  <div style="font-size:11px; margin-top:6px;">
                    {podcast_footer}
                  </div>

                </td>
              </tr>
            </table>
          </td>
        </tr>
      </table>

      <!-- INDIA MARKET RECAP -->
      <table width="100%" cellpadding="0" cellspacing="0" border="0" style="margin-top:24px;">
        <tr>
          <td style="font-size:18px; font-weight:700; color:#0f172a; padding-bottom:12px;">
            {recap_heading}
          </td>
        </tr>
      </table>

      <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#ffffff; border-radius:12px; box-shadow:0 8px 20px rgba(15,23,42,0.06); margin-bottom:16px;">
        <tr>
          <td style="padding:16px 18px;">
            <p style="font-size:14px; line-height:1.6; margin:0 0 12px 0;">
              {recap_para1}
            </p>
            <p style="font-size:14px; line-height:1.6; margin:0 0 12px 0;">
              {recap_para2}
            </p>
            <p style="font-size:14px; line-height:1.6; margin:0 0 12px 0;">
              {recap_para3}
            </p>
            <p style="font-size:11px; color:#6b7280; margin:0;">
              {recap_source}
            </p>
          </td>
        </tr>
      </table>

      <!-- TRADING PLAYBOOK -->
      <table width="100%" cellpadding="0" cellspacing="0" border="0" style="margin-top:24px;">
        <tr>
          <td style="font-size:18px; font-weight:700; color:#0f172a; padding-bottom:12px;">
            {playbook_heading}
          </td>
        </tr>
      </table>

      <!-- FII/DII + RANGE CARD -->
      <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#ffffff; border-radius:12px; box-shadow:0 8px 20px rgba(15,23,42,0.06); margin-bottom:16px;">
        <tr>
          <td style="padding:16px 18px;">

            <table width="100%" cellpadding="0" cellspacing="0" border="0">
              <tr>
                <td style="font-size:14px; font-weight:600; color:#374151; padding-bottom:4px;">
                  {flows_heading}
                </td>
              </tr>
              <tr>
                <td style="font-size:11px; color:#6b7280; padding-bottom:10px;">
                  {flows_subtext}
                </td>
              </tr>
            </table>

            <!-- FII/DII TABLE -->
            <table width="100%" cellpadding="0" cellspacing="0" border="0" style="font-size:12px; border-collapse:collapse;">
              <thead>
                <tr>
                  <th align="left" style="padding:8px 6px; background:#eff6ff; border-bottom:1px solid #e5e7eb; font-size:11px; text-transform:uppercase; letter-spacing:0.04em; color:#4b5563;">Participant</th>
                  <th align="left" style="padding:8px 6px; background:#eff6ff; border-bottom:1px solid #e5e7eb; font-size:11px; text-transform:uppercase; letter-spacing:0.04em; color:#4b5563;">Previous Day (₹ Cr)</th>
                  <th align="left" style="padding:8px 6px; background:#eff6ff; border-bottom:1px solid #e5e7eb; font-size:11px; text-transform:uppercase; letter-spacing:0.04em; color:#4b5563;">MTD (₹ Cr)</th>
                  <th align="left" style="padding:8px 6px; background:#eff6ff; border-bottom:1px solid #e5e7eb; font-size:11px; text-transform:uppercase; letter-spacing:0.04em; color:#4b5563;">YTD (₹ Cr)</th>
                </tr>
              </thead>
              <tbody>
                <tr>
                  <td style="padding:8px 6px; border-bottom:1px solid #e5e7eb;">FII</td>
                  <td style="padding:8px 6px; border-bottom:1px solid #e5e7eb; color:{fii_prev_color}; font-weight:500;">{fii_prev}</td>
                  <td style="padding:8px 6px; border-bottom:1px solid #e5e7eb; color:{fii_mtd_color}; font-weight:500;">{fii_mtd}</td>
                  <td style="padding:8px 6px; border-bottom:1px solid #e5e7eb; color:{fii_ytd_color}; font-weight:500;">{fii_ytd}</td>
                </tr>
                <tr>
                  <td style="padding:8px 6px; border-bottom:1px solid #e5e7eb;">DII</td>
                  <td style="padding:8px 6px; border-bottom:1px solid #e5e7eb; color:{dii_prev_color}; font-weight:500;">{dii_prev}</td>
                  <td style="padding:8px 6px; border-bottom:1px solid #e5e7eb; color:{dii_mtd_color}; font-weight:500;">{dii_mtd}</td>
                  <td style="padding:8px 6px; border-bottom:1px solid #e5e7eb; color:{dii_ytd_color}; font-weight:500;">{dii_ytd}</td>
                </tr>
              </tbody>
            </table>

            <p style="font-size:11px; color:#6b7280; margin:8px 0 0 0;">
              {flows_source}
            </p>

            <!-- RANGE LEVELS -->
            <p style="font-size:11px; color:#6b7280; margin:12px 0 4px 0;">
              {range_heading}
            </p>

            <table width="100%" cellpadding="0" cellspacing="0" border="0" style="font-size:12px; border-collapse:collapse;">
              <thead>
                <tr>
                  <th align="left" style="padding:8px 6px; background:#eff6ff; border-bottom:1px solid #e5e7eb; font-size:11px; text-transform:uppercase; letter-spacing:0.04em; color:#4b5563;">Index</th>
                  <th align="left" style="padding:8px 6px; background:#eff6ff; border-bottom:1px solid #e5e7eb; font-size:11px; text-transform:uppercase; letter-spacing:0.04em; color:#4b5563;">Support 1</th>
                  <th align="left" style="padding:8px 6px; background:#eff6ff; border-bottom:1px solid #e5e7eb; font-size:11px; text-transform:uppercase; letter-spacing:0.04em; color:#4b5563;">Support 2</th>
                  <th align="left" style="padding:8px 6px; background:#eff6ff; border-bottom:1px solid #e5e7eb; font-size:11px; text-transform:uppercase; letter-spacing:0.04em; color:#4b5563;">Resistance 1</th>
                  <th align="left" style="padding:8px 6px; background:#eff6ff; border-bottom:1px solid #e5e7eb; font-size:11px; text-transform:uppercase; letter-spacing:0.04em; color:#4b5563;">Resistance 2</th>
                </tr>
              </thead>
              <tbody>
                <tr>
                  <td style="padding:8px 6px; border-bottom:1px solid #e5e7eb;">{range_row1_index}</td>
                  <td style="padding:8px 6px; border-bottom:1px solid #e5e7eb;">{range_row1_s1}</td>
                  <td style="padding:8px 6px; border-bottom:1px solid #e5e7eb;">{range_row1_s2}</td>
                  <td style="padding:8px 6px; border-bottom:1px solid #e5e7eb;">{range_row1_r1}</td>
                  <td style="padding:8px 6px; border-bottom:1px solid #e5e7eb;">{range_row1_r2}</td>
                </tr>
                <tr>
                  <td style="padding:8px 6px; border-bottom:1px solid #e5e7eb;">{range_row2_index}</td>
                  <td style="padding:8px 6px; border-bottom:1px solid #e5e7eb;">{range_row2_s1}</td>
                  <td style="padding:8px 6px; border-bottom:1px solid #e5e7eb;">{range_row2_s2}</td>
                  <td style="padding:8px 6px; border-bottom:1px solid #e5e7eb;">{range_row2_r1}</td>
                  <td style="padding:8px 6px; border-bottom:1px solid #e5e7eb;">{range_row2_r2}</td>
                </tr>
              </tbody>
            </table>

            <p style="font-size:11px; color:#6b7280; margin:8px 0 0 0;">
              {range_comment}
            </p>

          </td>
        </tr>
      </table>

      <!-- MARKET OUTLOOK CARD -->
      <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#ffffff; border-radius:12px; box-shadow:0 8px 20px rgba(15,23,42,0.06); margin-bottom:16px;">
        <tr>
          <td style="padding:16px 18px;">
            <p style="font-size:14px; font-weight:600; color:#374151; margin:0 0 10px 0;">
              {outlook_heading}
            </p>
            <p style="font-size:14px; line-height:1.6; margin:0 0 10px 0;">
              {outlook_para1}
            </p>
            <p style="font-size:14px; line-height:1.6; margin:0 0 10px 0;">
              {outlook_para2}
            </p>
            <p style="font-size:14px; line-height:1.6; margin:0;">
              {outlook_para3}
            </p>
          </td>
        </tr>
      </table>

      <!-- F&O TRADE SETUP -->
      <table width="100%" cellpadding="0" cellspacing="0" border="0" style="margin-top:24px;">
        <tr>
          <td style="font-size:18px; font-weight:700; color:#0f172a; padding-bottom:12px;">
            {fo_heading}
          </td>
        </tr>
      </table>

      <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#ffffff; border-radius:12px; box-shadow:0 8px 20px rgba(15,23,42,0.06); margin-bottom:16px;">
        <tr>
          <td style="padding:16px 18px;">
            <p style="font-size:14px; font-weight:600; margin:0 0 8px 0;">{nifty_option_heading}</p>
            <ul style="margin:0 0 12px 18px; padding:0; font-size:14px; line-height:1.5;">
              {nifty_option_points}
            </ul>

            <p style="font-size:14px; font-weight:600; margin:8px 0 8px 0;">{bank_option_heading}</p>
            <ul style="margin:0 0 12px 18px; padding:0; font-size:14px; line-height:1.5;">
              {bank_option_points}
            </ul>

            <p style="font-size:11px; color:#6b7280; margin:0 0 6px 0;">
              {fo_source}
            </p>
            <p style="font-size:11px; color:#6b7280; margin:0;">
              {fo_ban_stocks}
            </p>
          </td>
        </tr>
      </table>

      <!-- GLOBAL MARKET PULSE -->
      <table width="100%" cellpadding="0" cellspacing="0" border="0" style="margin-top:24px;">
        <tr>
          <td style="font-size:18px; font-weight:700; color:#0f172a; padding-bottom:12px;">
            {global_heading}
          </td>
        </tr>
      </table>

      <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#ffffff; border-radius:12px; box-shadow:0 8px 20px rgba(15,23,42,0.06); margin-bottom:16px;">
        <tr>
          <td style="padding:16px 18px;">
            <p style="font-size:14px; line-height:1.6; margin:0 0 10px 0;">
              {global_para1}
            </p>
            <p style="font-size:11px; color:#6b7280; margin:0;">
              {global_source}
            </p>
          </td>
        </tr>
      </table>

      <!-- SECTOR MOVERS -->
      <table width="100%" cellpadding="0" cellspacing="0" border="0" style="margin-top:24px;">
        <tr>
          <td style="font-size:18px; font-weight:700; color:#0f172a; padding-bottom:12px;">
            {sector_heading}
          </td>
        </tr>
      </table>

      <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#ffffff; border-radius:12px; box-shadow:0 8px 20px rgba(15,23,42,0.06); margin-bottom:16px;">
        <tr>
          <td style="padding:16px 18px;">
            <table width="100%" cellpadding="0" cellspacing="0" border="0" style="font-size:12px; border-collapse:collapse;">
              <thead>
                <tr>
                  <th align="left" style="padding:8px 6px; background:#eff6ff; border-bottom:1px solid #e5e7eb; font-size:11px; text-transform:uppercase; letter-spacing:0.04em; color:#4b5563;">Leaders</th>
                  <th align="left" style="padding:8px 6px; background:#eff6ff; border-bottom:1px solid #e5e7eb; font-size:11px; text-transform:uppercase; letter-spacing:0.04em; color:#4b5563;">Laggards</th>
                </tr>
              </thead>
              <tbody>
                <tr>
                  <td style="padding:8px 6px; border-bottom:1px solid #e5e7eb;">{sector_leaders}</td>
                  <td style="padding:8px 6px; border-bottom:1px solid #e5e7eb;">{sector_laggards}</td>
                </tr>
              </tbody>
            </table>
            <p style="font-size:11px; color:#6b7280; margin:8px 0 0 0;">
              {sector_source}
            </p>
          </td>
        </tr>
      </table>

      <!-- STOCKS IN FOCUS -->
      <table width="100%" cellpadding="0" cellspacing="0" border="0" style="margin-top:24px;">
        <tr>
          <td style="font-size:18px; font-weight:700; color:#0f172a; padding-bottom:12px;">
            {stocks_heading}
          </td>
        </tr>
      </table>

      <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#ffffff; border-radius:12px; box-shadow:0 8px 20px rgba(15,23,42,0.06); margin-bottom:16px;">
        <tr>
          <td style="padding:16px 18px;">
            <table width="100%" cellpadding="0" cellspacing="0" border="0" style="font-size:12px; border-collapse:collapse;">
              <thead>
                <tr>
                  <th align="left" style="padding:8px 6px; background:#eff6ff; border-bottom:1px solid #e5e7eb; font-size:11px; text-transform:uppercase; letter-spacing:0.04em; color:#4b5563;">Bucket</th>
                  <th align="left" style="padding:8px 6px; background:#eff6ff; border-bottom:1px solid #e5e7eb; font-size:11px; text-transform:uppercase; letter-spacing:0.04em; color:#4b5563;">Stock</th>
                  <th align="left" style="padding:8px 6px; background:#eff6ff; border-bottom:1px solid #e5e7eb; font-size:11px; text-transform:uppercase; letter-spacing:0.04em; color:#4b5563;">Price %</th>
                  <th align="left" style="padding:8px 6px; background:#eff6ff; border-bottom:1px solid #e5e7eb; font-size:11px; text-transform:uppercase; letter-spacing:0.04em; color:#4b5563;">OI %</th>
                  <th align="left" style="padding:8px 6px; background:#eff6ff; border-bottom:1px solid #e5e7eb; font-size:11px; text-transform:uppercase; letter-spacing:0.04em; color:#4b5563;">Interpretation</th>
                </tr>
              </thead>
              <tbody>
                {stocks_rows_html}
              </tbody>
            </table>
            <p style="font-size:11px; color:#6b7280; margin:8px 0 0 0;">
              {stocks_source}
            </p>
          </td>
        </tr>
      </table>

      <!-- KEY EVENTS -->
      <table width="100%" cellpadding="0" cellspacing="0" border="0" style="margin-top:24px;">
        <tr>
          <td style="font-size:18px; font-weight:700; color:#0f172a; padding-bottom:12px;">
            {events_heading}
          </td>
        </tr>
      </table>

      <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#ffffff; border-radius:12px; box-shadow:0 8px 20px rgba(15,23,42,0.06); margin-bottom:16px;">
        <tr>
          <td style="padding:16px 18px;">
            <table width="100%" cellpadding="0" cellspacing="0" border="0" style="font-size:12px; border-collapse:collapse;">
              <thead>
                <tr>
                  <th align="left" style="padding:8px 6px; background:#eff6ff; border-bottom:1px solid #e5e7eb; font-size:11px; text-transform:uppercase; letter-spacing:0.04em; color:#4b5563;">Date</th>
                  <th align="left" style="padding:8px 6px; background:#eff6ff; border-bottom:1px solid #e5e7eb; font-size:11px; text-transform:uppercase; letter-spacing:0.04em; color:#4b5563;">Country</th>
                  <th align="left" style="padding:8px 6px; background:#eff6ff; border-bottom:1px solid #e5e7eb; font-size:11px; text-transform:uppercase; letter-spacing:0.04em; color:#4b5563;">Event</th>
                </tr>
              </thead>
              <tbody>
                {events_rows_html}
              </tbody>
            </table>
            <p style="font-size:11px; color:#6b7280; margin:8px 0 0 0;">
              {events_source}
            </p>
          </td>
        </tr>
      </table>

      <!-- CORPORATE HIGHLIGHTS -->
      <table width="100%" cellpadding="0" cellspacing="0" border="0" style="margin-top:24px;">
        <tr>
          <td style="font-size:18px; font-weight:700; color:#0f172a; padding-bottom:12px;">
            {corp_heading}
          </td>
        </tr>
      </table>

      <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background:#ffffff; border-radius:12px; box-shadow:0 8px 20px rgba(15,23,42,0.06);">
        <tr>
          <td style="padding:16px 18px;">
            <table width="100%" cellpadding="0" cellspacing="0" border="0" style="font-size:12px; border-collapse:collapse;">
              <thead>
                <tr>
                  <th align="left" style="padding:8px 6px; background:#eff6ff; border-bottom:1px solid #e5e7eb; font-size:11px; text-transform:uppercase; letter-spacing:0.04em; color:#4b5563;">Company / Theme</th>
                  <th align="left" style="padding:8px 6px; background:#eff6ff; border-bottom:1px solid #e5e7eb; font-size:11px; text-transform:uppercase; letter-spacing:0.04em; color:#4b5563;">Update</th>
                </tr>
              </thead>
              <tbody>
                {corp_rows_html}
              </tbody>
            </table>
            <p style="font-size:11px; color:#6b7280; margin:8px 0 0 0;">
              {corp_source}
            </p>
          </td>
        </tr>
      </table>

      <!-- DISCLAIMER -->
      <table width="100%" cellpadding="0" cellspacing="0" border="0" style="margin-top:24px;">
                <tr>
                    <td colspan="2" style="padding: 0px; font-size:11px;">
                        <p style="margin: 0;"><strong> Note :</strong>This Morning Note is prepared using information from the Bajaj Broking “Morning Bell” dated 16 December 2025 and other publicly available sources. It is meant for informational purposes only and does not constitute investment advice or a recommendation to buy, sell or hold any security. Investments in securities markets are subject to market risks. Please read all related documents carefully before investing.</p><br>
                        <p style="margin: 0;"><strong> Disclaimer :</strong>Please do not reply to this email, as responses to this
                            address will not be received or monitored. For any queries, feel free to contact us at
                            connect@bajajbroking.in.</p>
                        <p style="margin: 0; line-height: 16px; font-size:11px;">Investments in the securities market are subject to market risk.
                            Read all related documents carefully before investing. REG OFFICE: Bajaj Auto Limited Complex, Mumbai
                            Pune Road, Akurdi, Pune 411035. Corp. Office: Bajaj Financial Securities Limited. 1st Floor, Mantri IT
                            Park, Tower B, Unit No. 9 & 10, Viman Nagar, Pune, Maharashtra 411014. SEBI Registration No.:
                            INZ000218931 | BSE Cash/F&O/CDS Member ID: 6706 | NSE Cash/F&O/CDS Member ID: 90177 | SEBI DP
                            Registration No.: IN-DP-418-2019 | CDSL DP No.: 12088600 | NSDL DP No.: IN304030 | AMFI RegistrationNo.:
                            ARN-163403 | SEBI Registration No. (Research Analyst/Entity): INH000010043 | Website:
                            https://www.bajajbroking.in </p>
            
                        <p style="margin: 0; font-size:11px;">
                            Compliance Officer: Mr. Boudhayan Ghosh (For Broking/DP/Research) | Email:
                            compliance_sec@bajajbroking.in | Contact No.: 020-4857 4486.
                        </p>
            
                        <p style="margin: 0; font-size:11px;">
                            For any queries/information, you can call the toll-free number 1800-833-8888 or write to us on
                            <a href="mailto:connect@bajajbroking.in"
                                style="color: #1414DF; font-weight: 450;">connect@bajajbroking.in</a>.
                            <br>
                            For any Investor Grievances, write to compliance on
                            <a href="mailto:compliance_sec@bajajbroking.in"
                                style="color: #1414DF; font-weight: 450;">compliance_sec@bajajbroking.in</a> or
                            <a href="mailto:compliance_dp@bajajbroking.in"
                                style="color: #1414DF; font-weight: 450;">compliance_dp@bajajbroking.in</a>
                        </p>
            
                        <p style="margin: 0; font-size:11px;">
                            Kindly refer to
                            <a href="https://www.bajajbroking.in/disclaimer"
                                style="color: #1414DF; font-weight: 450;">https://www.bajajbroking.in/disclaimer</a>
                            for detailed disclaimer and risk factors.
                        </p>
                    </td>
                </tr>
            </table>

          </td>
        </tr>
      </table>
      <!-- /PAGE WRAPPER -->

    </td>
  </tr>
</table>

</body>
</html>

"""

# ---------- Core generator using Excel ----------

def generate_html_from_excel(excel_bytes: bytes) -> str:
    wb = load_workbook(BytesIO(excel_bytes), data_only=True)
    ws = wb["Sheet1"]

    # ----- MARKET SNAPSHOT -----
    raw_gift_value   = cell(ws, "A6")
    raw_gift_change  = cell(ws, "A7")
    raw_nifty_value  = cell(ws, "B6")
    raw_nifty_change = cell(ws, "B7")
    raw_sensex_value = cell(ws, "C6")
    raw_sensex_change= cell(ws, "C7")

    raw_bank_value   = cell(ws, "A9")
    raw_bank_change  = cell(ws, "A10")
    raw_vix_value    = cell(ws, "B9")
    raw_vix_change   = cell(ws, "B10")
    raw_usdinr_value = cell(ws, "C9")
    raw_usdinr_change= cell(ws, "C10")

    gift_value   = fmt_number(raw_gift_value, 1)
    gift_change  = fmt_percent(raw_gift_change, 2)
    nifty_value  = fmt_number(raw_nifty_value, 2)
    nifty_change = fmt_percent(raw_nifty_change, 2)
    sensex_value = fmt_number(raw_sensex_value, 2)
    sensex_change= fmt_percent(raw_sensex_change, 2)

    bank_value   = fmt_number(raw_bank_value, 2)
    bank_change  = fmt_percent(raw_bank_change, 2)
    vix_value    = fmt_number(raw_vix_value, 2)
    vix_change   = fmt_percent(raw_vix_change, 2)
    usdinr_value = fmt_number(raw_usdinr_value, 4)
    usdinr_change= fmt_percent(raw_usdinr_change, 2)

    # ----- FII/DII + RANGE -----
    playbook_heading = cell(ws, "A19")
    flows_heading    = cell(ws, "A20")
    flows_subtext    = cell(ws, "A22")

    raw_fii_prev = cell(ws, "B24")
    raw_fii_mtd  = cell(ws, "C24")
    raw_fii_ytd  = cell(ws, "D24")

    raw_dii_prev = cell(ws, "B25")
    raw_dii_mtd  = cell(ws, "C25")
    raw_dii_ytd  = cell(ws, "D25")

    fii_prev = fmt_number(raw_fii_prev, 0)
    fii_mtd  = fmt_number(raw_fii_mtd, 2)
    fii_ytd  = fmt_number(raw_fii_ytd, 2)

    dii_prev = fmt_number(raw_dii_prev, 0)
    dii_mtd  = fmt_number(raw_dii_mtd, 2)
    dii_ytd  = fmt_number(raw_dii_ytd, 2)

    flows_source = cell(ws, "A26")

    # ----- F&O points -----
    nifty_points_html = ""
    for row in range(45, 48):
        txt = cell(ws, f"A{row}")
        if txt:
            nifty_points_html += f"<li>{txt}</li>"

    # ----- STOCKS IN FOCUS -----
    stocks_rows_html = ""
    for r in range(73, 89):
        bucket = cell(ws, f"A{r}")
        stock  = cell(ws, f"B{r}")
        raw_price = cell(ws, f"C{r}")
        raw_oi    = cell(ws, f"D{r}")
        interp    = cell(ws, f"E{r}")

        if not (bucket or stock or raw_price or raw_oi or interp):
            break

        price = fmt_percent(raw_price, 2, show_sign=False)
        oi    = fmt_percent(raw_oi, 2, show_sign=False)

        stocks_rows_html += f"""
      <tr>
        <td style="padding:8px 6px; border-bottom:1px solid #e5e7eb;">{bucket}</td>
        <td style="padding:8px 6px; border-bottom:1px solid #e5e7eb;">{stock}</td>
        <td style="padding:8px 6px; border-bottom:1px solid #e5e7eb;">{price}</td>
        <td style="padding:8px 6px; border-bottom:1px solid #e5e7eb;">{oi}</td>
        <td style="padding:8px 6px; border-bottom:1px solid #e5e7eb;">{interp}</td>
      </tr>"""

    # ----- KEY EVENTS -----
    events_rows_html = ""
    for r in range(94, 111):
        raw_date  = ws[f"A{r}"].value
        ev_date   = fmt_date(raw_date)
        ev_country = cell(ws, f"B{r}")
        ev_event   = cell(ws, f"C{r}")

        if not (ev_date or ev_country or ev_event):
            break

        events_rows_html += f"""
      <tr>
        <td style="padding:8px 6px; border-bottom:1px solid #e5e7eb;">{ev_date}</td>
        <td style="padding:8px 6px; border-bottom:1px solid #e5e7eb;">{ev_country}</td>
        <td style="padding:8px 6px; border-bottom:1px solid #e5e7eb;">{ev_event}</td>
      </tr>"""

    # ----- CORPORATE HIGHLIGHTS -----
    corp_rows_html = ""
    for r in range(105, 115):
        company = cell(ws, f"A{r}")
        update  = cell(ws, f"B{r}")
        if not (company or update):
            break
        corp_rows_html += f"""
      <tr>
        <td style="padding:8px 6px; border-bottom:1px solid #e5e7eb;"><strong>{company}</strong></td>
        <td style="padding:8px 6px; border-bottom:1px solid #e5e7eb;">{update}</td>
      </tr>"""

    # ----- CONTEXT (same as earlier script) -----
    context = {
        "title":            cell(ws, "A3"),
        "date_line":        cell(ws, "A2"),
        "main_heading":     cell(ws, "A3"),

        "market_snapshot_heading": cell(ws, "A4"),
        "gift_label":       cell(ws, "A5"),
        "nifty_label":      cell(ws, "B5"),
        "sensex_label":     cell(ws, "C5"),
        "bank_label":       cell(ws, "A8"),
        "vix_label":        cell(ws, "B8"),
        "usdinr_label":     cell(ws, "C8"),

        "gift_value":   gift_value,
        "gift_change":  gift_change,
        "gift_color":   perc_color(gift_change),

        "nifty_value":  nifty_value,
        "nifty_change": nifty_change,
        "nifty_color":  perc_color(nifty_change),

        "sensex_value": sensex_value,
        "sensex_change":sensex_change,
        "sensex_color": perc_color(sensex_change),

        "bank_value":   bank_value,
        "bank_change":  bank_change,
        "bank_color":   perc_color(bank_change),

        "vix_value":    vix_value,
        "vix_change":   vix_change,
        "vix_color":    perc_color(vix_change),

        "usdinr_value": usdinr_value,
        "usdinr_change":usdinr_change,
        "usdinr_color": perc_color(usdinr_change),

        "podcast_tagline":       cell(ws, "D5"),
        "podcast_para1":         cell(ws, "D6"),
        "podcast_para2":         cell(ws, "D8"),
        "podcast_link":          cell(ws, "B16"),
        "podcast_button_label":  cell(ws, "B17"),
        "podcast_footer":        cell(ws, "D11"),

        "recap_heading":   cell(ws, "A13"),
        "recap_para1":     cell(ws, "A14"),
        "recap_para2":     cell(ws, "A15"),
        "recap_para3":     cell(ws, "A16"),
        "recap_source":    cell(ws, "A17"),

        "playbook_heading": playbook_heading,
        "flows_heading":    flows_heading,
        "flows_subtext":    flows_subtext,

        "fii_prev":        fii_prev,
        "fii_prev_color":  perc_color(fii_prev),
        "fii_mtd":         fii_mtd,
        "fii_mtd_color":   perc_color(fii_mtd),
        "fii_ytd":         fii_ytd,
        "fii_ytd_color":   perc_color(fii_ytd),

        "dii_prev":        dii_prev,
        "dii_prev_color":  perc_color(dii_prev),
        "dii_mtd":         dii_mtd,
        "dii_mtd_color":   perc_color(dii_mtd),
        "dii_ytd":         dii_ytd,
        "dii_ytd_color":   perc_color(dii_ytd),
        "flows_source":    flows_source,

        "range_heading":   cell(ws, "A27"),
        "range_row1_index": cell(ws, "A30"),
        "range_row1_s1":    fmt_number(cell(ws, "B30"), 0),
        "range_row1_s2":    fmt_number(cell(ws, "C30"), 0),
        "range_row1_r1":    fmt_number(cell(ws, "D30"), 0),
        "range_row1_r2":    fmt_number(cell(ws, "E30"), 0),

        "range_row2_index": cell(ws, "A31"),
        "range_row2_s1":    fmt_number(cell(ws, "B31"), 0),
        "range_row2_s2":    fmt_number(cell(ws, "C31"), 0),
        "range_row2_r1":    fmt_number(cell(ws, "D31"), 0),
        "range_row2_r2":    fmt_number(cell(ws, "E31"), 0),

        "range_comment":   cell(ws, "A32"),

        "outlook_heading": cell(ws, "A35"),
        "outlook_para1":   cell(ws, "A37"),
        "outlook_para2":   cell(ws, "A39"),
        "outlook_para3":   cell(ws, "A41"),

        "fo_heading":          cell(ws, "A43"),
        "nifty_option_heading": cell(ws, "A44"),

        "nifty_option_points" : nifty_points_html,
        "bank_option_heading":  cell(ws, "A50"),
        "bank_option_points":   nifty_points_html,
        "fo_source":            cell(ws, "A56"),
        "fo_ban_stocks":        cell(ws, "A57"),

        "global_heading": cell(ws, "A59"),
        "global_para1":   cell(ws, "A61"),
        "global_source":  cell(ws, "A63"),

        "sector_heading": cell(ws, "A65"),
        "sector_leaders": cell(ws, "A67"),
        "sector_laggards":cell(ws, "E67"),
        "sector_source":  cell(ws, "A69"),

        "stocks_heading":   cell(ws, "A71"),
        "stocks_rows_html": stocks_rows_html,
        "stocks_source":    cell(ws, "A90"),

        "events_heading":   cell(ws, "A92"),
        "events_rows_html": events_rows_html,
        "events_source":    cell(ws, "A101"),

        "corp_heading":   cell(ws, "A103"),
        "corp_rows_html": corp_rows_html,
        "corp_source":    cell(ws, "A116"),

        "note_text":           cell(ws, "A118"),
        "disclaimer_main":     cell(ws, "B76"),
        "disclaimer_reg":      cell(ws, "B77"),
        "compliance_contact":  cell(ws, "B78"),
        "support_contact":     cell(ws, "B79"),
        "disclaimer_link_text":cell(ws, "B80"),
    }

    wb.close()
    return HTML_TEMPLATE.format(**context)

# ---------- FastAPI endpoint ----------

@app.post("/generate-html")
async def generate_html(file: UploadFile = File(...)):
    """
    Upload Excel, get Morning Note HTML.
    """
    excel_bytes = await file.read()
    html_str = generate_html_from_excel(excel_bytes)

    buf = BytesIO(html_str.encode("utf-8"))
    filename = f"MorningNote_{date.today():%d-%b-%Y}.html"

    return StreamingResponse(
        buf,
        media_type="text/html",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
